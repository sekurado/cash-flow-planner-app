from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]
from pydantic import BaseModel, Field, ValidationError, field_validator

from src.domain.date_pattern import parse_pattern
from src.domain.entities import EntryCreateDTO, EntryType, ImportResult, ImportRowError
from src.domain.exceptions import DatePatternParseError

_REQUIRED_FIELDS = ("name", "date_pattern", "amount", "currency", "type")
_OPTIONAL_FIELDS = ("category",)
_ALL_FIELDS = _REQUIRED_FIELDS + _OPTIONAL_FIELDS


class ImportRowSchema(BaseModel):
    name: str
    date_pattern: str
    amount: float
    currency: str
    entry_type: EntryType = Field(alias="type")
    category: str | None = None

    @field_validator("name")
    @classmethod
    def _strip_name(cls, value: str) -> str:
        stripped = value.strip()
        if not stripped:
            msg = "Name is required"
            raise ValueError(msg)
        return stripped

    @field_validator("currency")
    @classmethod
    def _strip_currency(cls, value: str) -> str:
        stripped = value.strip()
        if not stripped:
            msg = "Currency is required"
            raise ValueError(msg)
        return stripped.upper()

    @field_validator("date_pattern")
    @classmethod
    def _validate_date_pattern(cls, value: str) -> str:
        stripped = value.strip()
        try:
            parse_pattern(stripped)
        except DatePatternParseError as exc:
            raise ValueError(str(exc)) from exc
        return stripped

    @field_validator("entry_type", mode="before")
    @classmethod
    def _normalize_entry_type(cls, value: object) -> object:
        if isinstance(value, str):
            return value.strip().lower()
        return value


class ImportService:
    """Parses CSV and Excel entry files into validated import DTOs."""

    def read_headers(self, path: Path | str) -> list[str]:
        file_path = Path(path)
        suffix = file_path.suffix.lower()
        if suffix == ".csv":
            return self._read_csv_headers(file_path)
        if suffix == ".xlsx":
            return self._read_xlsx_headers(file_path)
        msg = f"Unsupported file type: {suffix}"
        raise ValueError(msg)

    def parse(self, path: Path | str, column_mapping: dict[str, str]) -> ImportResult:
        file_path = Path(path)
        suffix = file_path.suffix.lower()
        if suffix == ".csv":
            raw_rows = self._read_csv_rows(file_path)
        elif suffix == ".xlsx":
            raw_rows = self._read_xlsx_rows(file_path)
        else:
            msg = f"Unsupported file type: {suffix}"
            raise ValueError(msg)

        headers = list(raw_rows[0][1].keys()) if raw_rows else self.read_headers(file_path)
        mapping = self._resolve_mapping(headers, column_mapping)
        missing = [field for field in _REQUIRED_FIELDS if not mapping.get(field)]
        if missing:
            missing_label = ", ".join(missing)
            error_message = f"Missing required column: {missing_label}"
            return ImportResult(
                valid_rows=[],
                errors=[
                    ImportRowError(row_number=row_number, error_message=error_message)
                    for row_number, _ in raw_rows
                ],
            )

        valid_rows: list[EntryCreateDTO] = []
        errors: list[ImportRowError] = []
        for row_number, raw_row in raw_rows:
            if self._is_blank_row(raw_row):
                continue
            mapped = self._apply_mapping(raw_row, mapping)
            try:
                validated = ImportRowSchema.model_validate(mapped)
            except ValidationError as exc:
                errors.append(
                    ImportRowError(
                        row_number=row_number,
                        error_message=self._format_validation_error(exc),
                    )
                )
                continue
            valid_rows.append(
                EntryCreateDTO(
                    entry_type=validated.entry_type,
                    name=validated.name,
                    date_pattern=validated.date_pattern,
                    amount=validated.amount,
                    currency=validated.currency,
                    category=validated.category,
                )
            )

        return ImportResult(valid_rows=valid_rows, errors=errors)

    def read_mapped_preview(
        self,
        path: Path | str,
        column_mapping: dict[str, str],
        *,
        limit: int = 5,
    ) -> list[dict[str, str]]:
        """Return the first *limit* non-blank rows with column mapping applied."""
        file_path = Path(path)
        suffix = file_path.suffix.lower()
        if suffix == ".csv":
            raw_rows = self._read_csv_rows(file_path)
        elif suffix == ".xlsx":
            raw_rows = self._read_xlsx_rows(file_path)
        else:
            msg = f"Unsupported file type: {suffix}"
            raise ValueError(msg)

        headers = list(raw_rows[0][1].keys()) if raw_rows else self.read_headers(file_path)
        mapping = self._resolve_mapping(headers, column_mapping)
        preview: list[dict[str, str]] = []
        for _, raw_row in raw_rows:
            if self._is_blank_row(raw_row):
                continue
            preview.append(self._apply_mapping(raw_row, mapping))
            if len(preview) >= limit:
                break
        return preview

    def _read_csv_headers(self, path: Path) -> list[str]:
        with path.open(newline="", encoding="utf-8-sig") as handle:
            sample = handle.read(8192)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample)
            except csv.Error:
                dialect = csv.excel
            reader = csv.DictReader(handle, dialect=dialect)
            return list(reader.fieldnames or [])

    def _read_csv_rows(self, path: Path) -> list[tuple[int, dict[str, str]]]:
        with path.open(newline="", encoding="utf-8-sig") as handle:
            sample = handle.read(8192)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample)
            except csv.Error:
                dialect = csv.excel
            reader = csv.DictReader(handle, dialect=dialect)
            rows: list[tuple[int, dict[str, str]]] = []
            for row_number, row in enumerate(reader, start=2):
                normalized = {
                    key: self._cell_to_str(value) for key, value in row.items() if key is not None
                }
                rows.append((row_number, normalized))
            return rows

    def _read_xlsx_headers(self, path: Path) -> list[str]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.worksheets[0]
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if header_row is None:
                return []
            return [self._cell_to_str(cell) for cell in header_row if self._cell_to_str(cell)]
        finally:
            workbook.close()

    def _read_xlsx_rows(self, path: Path) -> list[tuple[int, dict[str, str]]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.worksheets[0]
            row_iter = sheet.iter_rows(values_only=True)
            header_cells = next(row_iter, None)
            if header_cells is None:
                return []
            headers = [self._cell_to_str(cell) for cell in header_cells]
            rows: list[tuple[int, dict[str, str]]] = []
            for row_number, values in enumerate(row_iter, start=2):
                row_dict = {
                    header: self._cell_to_str(value)
                    for header, value in zip(headers, values, strict=False)
                    if header
                }
                rows.append((row_number, row_dict))
            return rows
        finally:
            workbook.close()

    def _resolve_mapping(
        self,
        headers: list[str],
        user_mapping: dict[str, str],
    ) -> dict[str, str]:
        header_lookup = {header.strip().lower(): header for header in headers if header.strip()}
        mapping = dict(user_mapping)
        for field in _ALL_FIELDS:
            if mapping.get(field):
                continue
            header = header_lookup.get(field.lower())
            if header is not None:
                mapping[field] = header
        return mapping

    def _apply_mapping(
        self,
        raw_row: dict[str, str],
        mapping: dict[str, str],
    ) -> dict[str, str]:
        mapped: dict[str, str] = {}
        for field, header in mapping.items():
            if not header:
                continue
            if header not in raw_row:
                continue
            mapped[field] = raw_row[header]
        return mapped

    def _is_blank_row(self, row: dict[str, str]) -> bool:
        return not any(value.strip() for value in row.values())

    def _cell_to_str(self, value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, bool):
            return str(value)
        if isinstance(value, int):
            return str(value)
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return str(value)
        return str(value).strip()

    def _format_validation_error(self, exc: ValidationError) -> str:
        messages = [error["msg"] for error in exc.errors()]
        return "; ".join(messages)
