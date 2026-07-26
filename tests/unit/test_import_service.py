from __future__ import annotations

import csv
from pathlib import Path

import pytest
from openpyxl import Workbook

from src.domain.entities import EntryType
from src.integrations.import_service import ImportService

_HEADERS = ("name", "date_pattern", "amount", "currency", "type")


def _write_csv(
    path: Path, rows: list[dict[str, str]], *, headers: tuple[str, ...] = _HEADERS
) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(headers))
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _valid_row(
    *,
    name: str = "Salary",
    date_pattern: str = "10..",
    amount: str = "5000",
    currency: str = "USD",
    entry_type: str = "income",
) -> dict[str, str]:
    return {
        "name": name,
        "date_pattern": date_pattern,
        "amount": amount,
        "currency": currency,
        "type": entry_type,
    }


@pytest.fixture
def import_service() -> ImportService:
    return ImportService()


@pytest.mark.unit
def test_parse_valid_csv_returns_three_rows(
    import_service: ImportService,
    tmp_path: Path,
) -> None:
    csv_path = tmp_path / "entries.csv"
    _write_csv(
        csv_path,
        [
            _valid_row(name="Salary"),
            _valid_row(name="Rent", date_pattern="1..", amount="1200", entry_type="expense"),
            _valid_row(name="Bonus", date_pattern="15..", amount="250"),
        ],
    )

    result = import_service.parse(csv_path, {})

    assert len(result.valid_rows) == 3
    assert result.errors == []
    assert result.valid_rows[0].name == "Salary"
    assert result.valid_rows[0].entry_type == EntryType.INCOME
    assert result.valid_rows[1].entry_type == EntryType.EXPENSE


@pytest.mark.unit
def test_parse_invalid_date_pattern_splits_valid_and_error_rows(
    import_service: ImportService,
    tmp_path: Path,
) -> None:
    csv_path = tmp_path / "entries.csv"
    _write_csv(
        csv_path,
        [
            _valid_row(name="Salary"),
            _valid_row(name="Bad row", date_pattern="not-valid"),
            _valid_row(name="Rent", date_pattern="1..", amount="1200", entry_type="expense"),
        ],
    )

    result = import_service.parse(csv_path, {})

    assert len(result.valid_rows) == 2
    assert len(result.errors) == 1
    assert result.errors[0].row_number == 3


@pytest.mark.unit
def test_parse_column_mapping_with_non_default_headers(
    import_service: ImportService,
    tmp_path: Path,
) -> None:
    headers = ("Title", "Pattern", "Value", "Curr", "Kind")
    csv_path = tmp_path / "entries.csv"
    _write_csv(
        csv_path,
        [
            {
                "Title": "Salary",
                "Pattern": "10..",
                "Value": "5000",
                "Curr": "USD",
                "Kind": "income",
            },
            {
                "Title": "Rent",
                "Pattern": "1..",
                "Value": "1200",
                "Curr": "USD",
                "Kind": "expense",
            },
        ],
        headers=headers,
    )
    mapping = {
        "name": "Title",
        "date_pattern": "Pattern",
        "amount": "Value",
        "currency": "Curr",
        "type": "Kind",
    }

    result = import_service.parse(csv_path, mapping)

    assert len(result.valid_rows) == 2
    assert result.errors == []
    assert result.valid_rows[0].name == "Salary"
    assert result.valid_rows[1].name == "Rent"


@pytest.mark.unit
def test_parse_empty_file_returns_no_rows_or_errors(
    import_service: ImportService,
    tmp_path: Path,
) -> None:
    csv_path = tmp_path / "empty.csv"
    csv_path.write_text("", encoding="utf-8")

    result = import_service.parse(csv_path, {})

    assert result.valid_rows == []
    assert result.errors == []


@pytest.mark.unit
def test_parse_xlsx_imports_rows(
    import_service: ImportService,
    tmp_path: Path,
) -> None:
    xlsx_path = tmp_path / "entries.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(list(_HEADERS))
    sheet.append(["Salary", "10..", 5000, "USD", "income"])
    sheet.append(["Rent", "1..", 1200, "USD", "expense"])
    workbook.save(xlsx_path)

    result = import_service.parse(xlsx_path, {})

    assert len(result.valid_rows) == 2
    assert result.errors == []
    assert result.valid_rows[0].amount == 5000.0
    assert result.valid_rows[1].entry_type == EntryType.EXPENSE


@pytest.mark.unit
def test_parse_missing_required_column_puts_rows_in_errors(
    import_service: ImportService,
    tmp_path: Path,
) -> None:
    headers = ("name", "date_pattern", "currency", "type")
    csv_path = tmp_path / "entries.csv"
    _write_csv(
        csv_path,
        [
            {
                "name": "Salary",
                "date_pattern": "10..",
                "currency": "USD",
                "type": "income",
            },
            {
                "name": "Rent",
                "date_pattern": "1..",
                "currency": "USD",
                "type": "expense",
            },
        ],
        headers=headers,
    )

    result = import_service.parse(csv_path, {})

    assert result.valid_rows == []
    assert len(result.errors) == 2
    assert all("amount" in error.error_message.lower() for error in result.errors)


@pytest.mark.unit
def test_read_headers_reads_csv_and_xlsx_columns(
    import_service: ImportService,
    tmp_path: Path,
) -> None:
    csv_path = tmp_path / "entries.csv"
    _write_csv(csv_path, [_valid_row(name="Salary")])

    xlsx_path = tmp_path / "entries.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(list(_HEADERS))
    workbook.save(xlsx_path)

    assert import_service.read_headers(csv_path) == list(_HEADERS)
    assert import_service.read_headers(xlsx_path) == list(_HEADERS)


@pytest.mark.unit
def test_unsupported_file_type_raises_value_error(
    import_service: ImportService,
    tmp_path: Path,
) -> None:
    bad_path = tmp_path / "entries.txt"
    bad_path.write_text("name\nSalary", encoding="utf-8")

    with pytest.raises(ValueError, match="Unsupported file type"):
        import_service.read_headers(bad_path)

    with pytest.raises(ValueError, match="Unsupported file type"):
        import_service.parse(bad_path, {})


@pytest.mark.unit
def test_read_mapped_preview_returns_limited_mapped_rows(
    import_service: ImportService,
    tmp_path: Path,
) -> None:
    csv_path = tmp_path / "entries.csv"
    _write_csv(
        csv_path,
        [_valid_row(name=f"Entry {index}") for index in range(5)],
    )

    preview = import_service.read_mapped_preview(csv_path, {}, limit=2)

    assert len(preview) == 2
    assert preview[0]["name"] == "Entry 0"
    assert preview[1]["name"] == "Entry 1"


@pytest.mark.unit
def test_parse_skips_blank_rows(
    import_service: ImportService,
    tmp_path: Path,
) -> None:
    csv_path = tmp_path / "entries.csv"
    _write_csv(
        csv_path,
        [
            _valid_row(name="Salary"),
            {"name": "", "date_pattern": "", "amount": "", "currency": "", "type": ""},
            _valid_row(name="Rent", date_pattern="1..", amount="1200", entry_type="expense"),
        ],
    )

    result = import_service.parse(csv_path, {})

    assert len(result.valid_rows) == 2
    assert result.errors == []


@pytest.mark.unit
def test_parse_validation_errors_for_empty_name_and_currency(
    import_service: ImportService,
    tmp_path: Path,
) -> None:
    csv_path = tmp_path / "entries.csv"
    _write_csv(
        csv_path,
        [
            _valid_row(name="   "),
            _valid_row(name="Rent", currency="  "),
        ],
    )

    result = import_service.parse(csv_path, {})

    assert result.valid_rows == []
    assert len(result.errors) == 2
    assert "Name is required" in result.errors[0].error_message
    assert "Currency is required" in result.errors[1].error_message


@pytest.mark.unit
def test_parse_xlsx_handles_empty_sheet_and_non_integer_amounts(
    import_service: ImportService,
    tmp_path: Path,
) -> None:
    empty_path = tmp_path / "empty.xlsx"
    empty_workbook = Workbook()
    empty_workbook.active.delete_rows(1, 1)
    empty_workbook.save(empty_path)

    assert import_service.read_headers(empty_path) == []
    empty_result = import_service.parse(empty_path, {})
    assert empty_result.valid_rows == []
    assert empty_result.errors == []

    xlsx_path = tmp_path / "entries.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(list(_HEADERS))
    sheet.append(["Salary", "10..", 123.45, "usd", "income"])
    workbook.save(xlsx_path)

    result = import_service.parse(xlsx_path, {})

    assert len(result.valid_rows) == 1
    assert result.valid_rows[0].amount == 123.45
    assert result.valid_rows[0].currency == "USD"
